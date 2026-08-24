from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    send_file
)

import sqlite3
import uuid
import qrcode
import os

from datetime import datetime

from openpyxl import Workbook, load_workbook


app = Flask(__name__)

app.secret_key = "qr_attendance_secret_key"

DATABASE = "database.db"

EXCEL_FOLDER = "attendance_excel"


# =========================================================
# DATABASE
# =========================================================

def get_db_connection():

    connection = sqlite3.connect(DATABASE)

    connection.row_factory = sqlite3.Row

    return connection


def create_database():

    connection = get_db_connection()

    cursor = connection.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS students (

            id INTEGER PRIMARY KEY AUTOINCREMENT,

            roll_number TEXT UNIQUE NOT NULL,

            name TEXT NOT NULL,

            department TEXT NOT NULL

        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS attendance_sessions (

            id INTEGER PRIMARY KEY AUTOINCREMENT,

            session_id TEXT UNIQUE NOT NULL,

            subject_name TEXT NOT NULL,

            lecture_number INTEGER NOT NULL,

            lecture_time TEXT NOT NULL,

            student_limit INTEGER NOT NULL,

            attendance_count INTEGER DEFAULT 0,

            status TEXT DEFAULT 'OPEN'

        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS attendance (

            id INTEGER PRIMARY KEY AUTOINCREMENT,

            session_id TEXT NOT NULL,

            student_id INTEGER NOT NULL,

            roll_number TEXT NOT NULL,

            student_name TEXT NOT NULL,

            department TEXT NOT NULL,

            subject_name TEXT NOT NULL,

            lecture_number INTEGER NOT NULL,

            lecture_time TEXT NOT NULL,

            attendance_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP,

            UNIQUE(session_id, student_id)

        )
    """)

    connection.commit()

    connection.close()


# =========================================================
# EXCEL
# =========================================================

def get_excel_file_path(session):

    os.makedirs(
        EXCEL_FOLDER,
        exist_ok=True
    )

    safe_subject = "".join(

        character

        for character in session["subject_name"]

        if character.isalnum()
        or character in (" ", "_", "-")

    )

    safe_subject = safe_subject.strip()

    if not safe_subject:

        safe_subject = "Attendance"

    filename = (
        f"{safe_subject}_Lecture_"
        f"{session['lecture_number']}_"
        f"{session['session_id']}.xlsx"
    )

    return os.path.join(
        EXCEL_FOLDER,
        filename
    )


def create_excel_file(session):

    filepath = get_excel_file_path(session)

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "Attendance"

    headers = [
        "Date",
        "Attendance Time",
        "Subject",
        "Lecture Number",
        "Lecture Time",
        "Roll Number",
        "Student Name",
        "Department",
        "Status"
    ]

    worksheet.append(headers)

    workbook.save(filepath)

    return filepath


def add_attendance_to_excel(session, student):

    filepath = get_excel_file_path(session)

    if not os.path.exists(filepath):

        create_excel_file(session)

    workbook = load_workbook(filepath)

    worksheet = workbook.active

    now = datetime.now()

    current_date = now.strftime(
        "%d-%m-%Y"
    )

    current_time = now.strftime(
        "%H:%M:%S"
    )

    worksheet.append([

        current_date,

        current_time,

        session["subject_name"],

        session["lecture_number"],

        session["lecture_time"],

        student["roll_number"],

        student["name"],

        student["department"],

        "Present"

    ])

    workbook.save(filepath)


# =========================================================
# HOME
# =========================================================

@app.route("/")
def home():

    return render_template(
        "index.html"
    )


# =========================================================
# TEACHER PAGE
# =========================================================

@app.route("/teacher")
def teacher():

    return render_template(
        "teacher.html"
    )


# =========================================================
# CREATE SESSION
# =========================================================

@app.route(
    "/create_session",
    methods=["POST"]
)
def create_session():

    subject_name = request.form[
        "subject_name"
    ].strip()

    lecture_number = request.form[
        "lecture_number"
    ].strip()

    lecture_time = request.form[
        "lecture_time"
    ].strip()

    student_limit = request.form[
        "student_limit"
    ].strip()

    if not subject_name:

        flash("Please enter subject name.")

        return redirect(
            url_for("teacher")
        )

    try:

        lecture_number = int(
            lecture_number
        )

        student_limit = int(
            student_limit
        )

    except ValueError:

        flash(
            "Lecture number and student limit "
            "must be numbers."
        )

        return redirect(
            url_for("teacher")
        )

    if lecture_number < 1:

        flash(
            "Lecture number must be at least 1."
        )

        return redirect(
            url_for("teacher")
        )

    if student_limit < 1:

        flash(
            "Student limit must be at least 1."
        )

        return redirect(
            url_for("teacher")
        )

    session_id = str(
        uuid.uuid4()
    )

    connection = get_db_connection()

    connection.execute(
        """
        INSERT INTO attendance_sessions
        (
            session_id,
            subject_name,
            lecture_number,
            lecture_time,
            student_limit
        )

        VALUES (?, ?, ?, ?, ?)
        """,

        (
            session_id,
            subject_name,
            lecture_number,
            lecture_time,
            student_limit
        )
    )

    connection.commit()

    session = connection.execute(
        """
        SELECT *
        FROM attendance_sessions
        WHERE session_id = ?
        """,

        (session_id,)
    ).fetchone()

    connection.close()

    create_excel_file(session)

    qr_folder = os.path.join(
        "static",
        "qr_codes"
    )

    os.makedirs(
        qr_folder,
        exist_ok=True
    )

    qr_data = url_for(
        "student_attendance",
        session_id=session_id,
        _external=True
    )

    qr = qrcode.QRCode(

        version=1,

        error_correction=
        qrcode.constants.ERROR_CORRECT_H,

        box_size=10,

        border=4
    )

    qr.add_data(qr_data)

    qr.make(
        fit=True
    )

    qr_image = qr.make_image(
        fill_color="black",
        back_color="white"
    )

    qr_filename = (
        f"{session_id}.png"
    )

    qr_path = os.path.join(
        qr_folder,
        qr_filename
    )

    qr_image.save(
        qr_path
    )

    return render_template(
        "session_created.html",

        session_id=session_id,

        subject_name=subject_name,

        lecture_number=lecture_number,

        lecture_time=lecture_time,

        student_limit=student_limit,

        qr_filename=qr_filename
    )


# =========================================================
# TEACHER ATTENDANCE
# =========================================================

@app.route(
    "/teacher/attendance/<session_id>"
)
def teacher_attendance(session_id):

    connection = get_db_connection()

    session = connection.execute(
        """
        SELECT *
        FROM attendance_sessions
        WHERE session_id = ?
        """,

        (session_id,)
    ).fetchone()

    if session is None:

        connection.close()

        return render_template(
            "attendance_error.html",
            message="Session not found."
        )

    attendance_list = connection.execute(
        """
        SELECT *
        FROM attendance
        WHERE session_id = ?
        ORDER BY attendance_time ASC
        """,

        (session_id,)
    ).fetchall()

    connection.close()

    remaining = (
        session["student_limit"]
        - session["attendance_count"]
    )

    if remaining < 0:

        remaining = 0

    excel_path = get_excel_file_path(
        session
    )

    excel_exists = os.path.exists(
        excel_path
    )

    return render_template(
        "teacher_attendance.html",

        session=session,

        attendance_list=attendance_list,

        remaining=remaining,

        excel_exists=excel_exists
    )


# =========================================================
# DOWNLOAD EXCEL
# =========================================================

@app.route(
    "/download_excel/<session_id>"
)
def download_excel(session_id):

    connection = get_db_connection()

    session = connection.execute(
        """
        SELECT *
        FROM attendance_sessions
        WHERE session_id = ?
        """,

        (session_id,)
    ).fetchone()

    connection.close()

    if session is None:

        return "Session not found."

    excel_path = get_excel_file_path(
        session
    )

    if not os.path.exists(excel_path):

        return "Excel file not found."

    return send_file(
        excel_path,
        as_attachment=True
    )


# =========================================================
# STUDENT PAGE
# =========================================================

@app.route(
    "/attendance/<session_id>"
)
def student_attendance(session_id):

    connection = get_db_connection()

    session = connection.execute(
        """
        SELECT *
        FROM attendance_sessions
        WHERE session_id = ?
        """,

        (session_id,)
    ).fetchone()

    connection.close()

    if session is None:

        return render_template(
            "attendance_error.html",
            message="Session not found."
        )

    if session["status"] != "OPEN":

        return render_template(
            "attendance_closed.html",
            session=session
        )

    return render_template(
        "student.html",
        session=session
    )


# =========================================================
# SUBMIT ATTENDANCE
# =========================================================

@app.route(
    "/submit_attendance",
    methods=["POST"]
)
def submit_attendance():

    session_id = request.form[
        "session_id"
    ].strip()

    name = request.form[
        "name"
    ].strip()

    roll_number = request.form[
        "roll_number"
    ].strip()

    department = request.form[
        "department"
    ].strip()

    if (
        not name
        or not roll_number
        or not department
    ):

        return render_template(
            "attendance_error.html",
            message="Please fill all fields."
        )

    connection = get_db_connection()

    session = connection.execute(
        """
        SELECT *
        FROM attendance_sessions
        WHERE session_id = ?
        """,

        (session_id,)
    ).fetchone()

    if session is None:

        connection.close()

        return render_template(
            "attendance_error.html",
            message="Session not found."
        )

    if session["status"] != "OPEN":

        connection.close()

        return render_template(
            "attendance_closed.html",
            session=session
        )

    if (
        session["attendance_count"]
        >=
        session["student_limit"]
    ):

        connection.execute(
            """
            UPDATE attendance_sessions
            SET status = 'CLOSED'
            WHERE session_id = ?
            """,

            (session_id,)
        )

        connection.commit()

        connection.close()

        return render_template(
            "attendance_closed.html",
            session=session
        )

    student = connection.execute(
        """
        SELECT *
        FROM students
        WHERE roll_number = ?
        """,

        (roll_number,)
    ).fetchone()

    if student is None:

        connection.close()

        return render_template(
            "attendance_error.html",
            message="Roll number not registered."
        )

    if (
        student["name"].lower()
        !=
        name.lower()
    ):

        connection.close()

        return render_template(
            "attendance_error.html",
            message="Name does not match."
        )

    if (
        student["department"].lower()
        !=
        department.lower()
    ):

        connection.close()

        return render_template(
            "attendance_error.html",
            message="Department does not match."
        )

    existing = connection.execute(
        """
        SELECT *
        FROM attendance

        WHERE session_id = ?

        AND student_id = ?
        """,

        (
            session_id,
            student["id"]
        )
    ).fetchone()

    if existing is not None:

        connection.close()

        return render_template(
            "attendance_error.html",
            message="Attendance already marked."
        )

    try:

        connection.execute(
            """
            INSERT INTO attendance
            (
                session_id,
                student_id,
                roll_number,
                student_name,
                department,
                subject_name,
                lecture_number,
                lecture_time
            )

            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,

            (
                session_id,
                student["id"],
                student["roll_number"],
                student["name"],
                student["department"],
                session["subject_name"],
                session["lecture_number"],
                session["lecture_time"]
            )
        )

        connection.execute(
            """
            UPDATE attendance_sessions

            SET attendance_count =
                attendance_count + 1

            WHERE session_id = ?
            """,

            (session_id,)
        )

        connection.commit()

    except sqlite3.IntegrityError:

        connection.close()

        return render_template(
            "attendance_error.html",
            message="Attendance could not be recorded."
        )

    add_attendance_to_excel(
        session,
        student
    )

    updated_session = connection.execute(
        """
        SELECT *
        FROM attendance_sessions
        WHERE session_id = ?
        """,

        (session_id,)
    ).fetchone()

    if (
        updated_session["attendance_count"]
        >=
        updated_session["student_limit"]
    ):

        connection.execute(
            """
            UPDATE attendance_sessions

            SET status = 'CLOSED'

            WHERE session_id = ?
            """,

            (session_id,)
        )

        connection.commit()

        updated_session = connection.execute(
            """
            SELECT *
            FROM attendance_sessions
            WHERE session_id = ?
            """,

            (session_id,)
        ).fetchone()

    connection.close()

    return render_template(
        "attendance_success.html",

        student_name=student["name"],

        roll_number=student["roll_number"],

        session=updated_session
    )


# =========================================================
# STUDENTS
# =========================================================

@app.route("/students")
def students():

    connection = get_db_connection()

    students_list = connection.execute(
        """
        SELECT *
        FROM students
        ORDER BY roll_number
        """
    ).fetchall()

    connection.close()

    return render_template(
        "students.html",
        students=students_list
    )


# =========================================================
# ADD STUDENT
# =========================================================

@app.route(
    "/add_student",
    methods=["POST"]
)
def add_student():

    name = request.form[
        "name"
    ].strip()

    roll_number = request.form[
        "roll_number"
    ].strip()

    department = request.form[
        "department"
    ].strip()

    if (
        not name
        or not roll_number
        or not department
    ):

        flash(
            "Please fill all fields."
        )

        return redirect(
            url_for("students")
        )

    connection = get_db_connection()

    try:

        connection.execute(
            """
            INSERT INTO students
            (
                roll_number,
                name,
                department
            )

            VALUES (?, ?, ?)
            """,

            (
                roll_number,
                name,
                department
            )
        )

        connection.commit()

        flash(
            "Student added successfully."
        )

    except sqlite3.IntegrityError:

        flash(
            "Roll number already exists."
        )

    finally:

        connection.close()

    return redirect(
        url_for("students")
    )


# =========================================================
# DELETE STUDENT
# =========================================================

@app.route(
    "/delete_student/<int:student_id>"
)
def delete_student(student_id):

    connection = get_db_connection()

    connection.execute(
        """
        DELETE FROM students
        WHERE id = ?
        """,

        (student_id,)
    )

    connection.commit()

    connection.close()

    flash(
        "Student deleted successfully."
    )

    return redirect(
        url_for("students")
    )


# =========================================================
# RUN
# =========================================================

if __name__ == "__main__":

    create_database()

    os.makedirs(
        EXCEL_FOLDER,
        exist_ok=True
    )

    os.makedirs(
        "static/qr_codes",
        exist_ok=True
    )

    app.run(
        host="0.0.0.0",
        port=5000,
        debug=True
    )